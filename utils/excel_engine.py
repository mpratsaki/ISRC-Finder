"""
utils/excel_engine.py

Zero-to-Excel catalog generator: builds the "Stay Independent Catalog"
workbook (title/role/writers/IPI/PRO/%rights/ISRC/notes) from a list of
Spotify tracks, enriched with Tidal contributor credits and IPI LIST matches.
"""

import io
import re
import unicodedata
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from utils.tidal_api import validate_isrc, fetch_tidal_contributors_by_isrc
from utils.github_fetcher import _build_contributor_rows


# --------------------------------------------------------------------------
# Excel Generator (Zero-to-Excel) - builds in-memory
# --------------------------------------------------------------------------
def generate_new_catalog(tracks, ipi_lookup=None, progress_callback=None):
    ipi_lookup = ipi_lookup or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stay Independent Catalog"

    report = {
        "filled": [],
        "health_warnings": [],
        "tidal_fallbacks": [],
        "ipi_matches": 0,
    }

    # --- Ορισμός Στυλ (Γραμματοσειρές, Στοιχίσεις, Χρώματα, Περιγράμματα) ---
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    top_alignment = Alignment(vertical="top", wrap_text=True)
    sum_font = Font(bold=True, color="000000")  # Έντονη γραφή για το άθροισμα

    # Πιο έντονο μαύρο περίγραμμα (medium style)
    black_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )

    # Γκρι γέμισμα για το διαχωριστικό κελί
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Νέα σειρά στηλών βάσει του πρότυπου Excel
    headers = ["TITLE", "ROLE", "WRITERS", "IPI", "PRO", "% RIGHTS", "ISRC", "NOTES"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = black_border

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 35  # TITLE
    ws.column_dimensions["B"].width = 15  # ROLE
    ws.column_dimensions["C"].width = 32  # WRITERS
    ws.column_dimensions["D"].width = 16  # IPI
    ws.column_dimensions["E"].width = 18  # PRO
    ws.column_dimensions["F"].width = 15  # % RIGHTS
    ws.column_dimensions["G"].width = 20  # ISRC
    ws.column_dimensions["H"].width = 55  # NOTES

    insert_at = 2
    total_tracks = len(tracks)

    for track_index, track in enumerate(tracks, start=1):
        title = track.get("name") or ""
        spotify_artists = track.get("artists") or []
        isrc = track.get("isrc")
        clean_isrc = str(isrc or "").replace("-", "").strip().upper()

        if progress_callback:
            progress_callback(track_index, total_tracks, title)

        notes = []
        tidal_names = []
        tidal_note = None
        should_try_tidal = bool(clean_isrc) and validate_isrc(clean_isrc)

        if clean_isrc and not validate_isrc(clean_isrc):
            report["health_warnings"].append((title, isrc))
            notes.append("ISRC format invalid")
        elif not clean_isrc:
            notes.append("Missing ISRC")

        if should_try_tidal:
            tidal_names, tidal_note = fetch_tidal_contributors_by_isrc(clean_isrc)

        if tidal_names:
            contributor_names = tidal_names
            contributor_source = "tidal"
        else:
            contributor_names = spotify_artists
            contributor_source = "spotify_fallback"

            if should_try_tidal:
                note = tidal_note or "Tidal credits not found — used Spotify artists as fallback"
                notes.append(note)
                report["tidal_fallbacks"].append(title)
            elif clean_isrc:
                notes.append("Tidal lookup skipped — used Spotify artists as fallback")
            else:
                notes.append("Tidal lookup skipped — used Spotify artists as fallback")

        contributor_rows = _build_contributor_rows(contributor_names, ipi_lookup)
        report["ipi_matches"] += sum(1 for row in contributor_rows if row["matched"])

        needed_rows = max(1, len(contributor_rows))
        notes_text = "; ".join(dict.fromkeys(note for note in notes if note))

        # Αποθηκεύουμε την αρχική και τελική γραμμή του τρέχοντος τραγουδιού για τη συνάρτηση SUM
        start_row = insert_at
        end_row = insert_at + needed_rows - 1

        for i in range(needed_rows):
            current_row = insert_at + i

            # Επαναλαμβάνουμε TITLE, ISRC και NOTES σε ΚΑΘΕ γραμμή του πλαισίου
            ws.cell(row=current_row, column=1).value = title

            if isrc:
                ws.cell(row=current_row, column=7).value = isrc  # Το ISRC πήγε στη στήλη 7 (G)
            if notes_text:
                ws.cell(row=current_row, column=8).value = notes_text  # Τα NOTES πήγαν στη στήλη 8 (H)

            if i < len(contributor_rows):
                contributor = contributor_rows[i]
                ws.cell(row=current_row, column=3).value = contributor["writer"]  # WRITERS στήλη 3 (C)

                if contributor["ipi"] is not None:
                    ipi_cell = ws.cell(row=current_row, column=4)  # IPI στήλη 4 (D)
                    ipi_cell.value = contributor["ipi"]
                    ipi_cell.number_format = "0"

                if contributor["pro"]:
                    ws.cell(row=current_row, column=5).value = contributor["pro"]  # PRO στήλη 5 (E)

            # Εφαρμογή του μαύρου περιγράμματος σε όλα τα 8 κελιά της τρέχουσας γραμμής
            for col_num in range(1, 9):
                cell = ws.cell(row=current_row, column=col_num)
                cell.alignment = top_alignment
                cell.border = black_border

        # --- Προσθήκη Γκρι Διαχωριστικής Γραμμής & Δυναμικού Υπολογισμού ---
        separator_row = insert_at + needed_rows
        for col_num in range(1, 9):
            cell = ws.cell(row=separator_row, column=col_num)
            cell.fill = gray_fill
            cell.border = black_border

            # Εισαγωγή της δυναμικής Excel Formula (=SUM) στο κελί της στήλης 6 (% RIGHTS)
            if col_num == 6:
                cell.value = f"=SUM(F{start_row}:F{end_row})"
                cell.font = sum_font
                cell.alignment = center_alignment

        report["filled"].append(
            {
                "title": title,
                "contributors": [row["writer"] for row in contributor_rows] if contributor_rows else [],
                "isrc": isrc,
                "source": contributor_source,
                "notes": notes_text,
            }
        )

        # Το επόμενο τραγούδι ξεκινά μετά το block των writers + 1 γραμμή για το διαχωριστικό
        insert_at += needed_rows + 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, report


# --------------------------------------------------------------------------
# Filename helper
# --------------------------------------------------------------------------
def make_catalog_filename(playlist_name):
    safe_name = re.sub(r"\s+", "_", str(playlist_name or "playlist").strip())
    safe_name = unicodedata.normalize("NFKD", safe_name).encode("ascii", "ignore").decode("ascii")
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "", safe_name).strip("_")
    if not safe_name:
        safe_name = "playlist"

    date_part = datetime.now().strftime("%Y%m%d")
    return f"Catalog_{safe_name}_{date_part}.xlsx"
