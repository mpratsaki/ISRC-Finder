"""
utils/github_fetcher.py

Private GitHub Contents API helpers plus the existing IPI LIST parser and
metadata-health utilities.

The generic ``fetch_private_file_bytes`` function is shared by the IPI LIST
and Label Copy template fetchers. Existing public IPI functions retain their
names, arguments, return values, and caller-visible failure behaviour.
"""

from __future__ import annotations

import base64
import io
import re
import time
import urllib.parse
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
import requests
import streamlit as st


GITHUB_API_VERSION = "2022-11-28"
GITHUB_CONTENTS_TIMEOUT_SECONDS = 20
GITHUB_CONTENTS_MAX_RETRIES = 3
PRIVATE_FILE_CACHE_TTL_SECONDS = 15 * 60
IPI_LIST_CACHE_TTL_SECONDS = PRIVATE_FILE_CACHE_TTL_SECONDS
GITHUB_USER_AGENT = "stay-independent-tool"


class PrivateFileFetchError(RuntimeError):
    """Structured internal error that callers may translate for their context."""

    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


# --------------------------------------------------------------------------
# Streamlit secret configuration
# --------------------------------------------------------------------------
def _secret_text(key: str, default: str = "") -> str:
    try:
        value = st.secrets.get(key, default)
    except Exception:
        value = default
    return str(value or "").strip()


def get_private_ipi_config():
    """
    Reads the private GitHub source for the IPI LIST ground-truth Excel.

    Required Streamlit secrets:
      IPI_GITHUB_OWNER  -> GitHub user/org that owns the private repository
      IPI_GITHUB_REPO   -> private repository name
      IPI_GITHUB_PATH   -> path to the xlsx file inside the repository
      IPI_GITHUB_TOKEN  -> fine-grained PAT with Contents: Read-only on that repo

    Optional Streamlit secret:
      IPI_GITHUB_REF    -> branch, tag, or commit SHA. Defaults to "main".
    """
    required_keys = [
        "IPI_GITHUB_OWNER",
        "IPI_GITHUB_REPO",
        "IPI_GITHUB_PATH",
        "IPI_GITHUB_TOKEN",
    ]
    missing = [key for key in required_keys if not _secret_text(key)]
    if missing:
        raise RuntimeError("Λείπουν Streamlit secrets για το IPI LIST: " + ", ".join(missing))

    return {
        "owner": _secret_text("IPI_GITHUB_OWNER"),
        "repo": _secret_text("IPI_GITHUB_REPO"),
        "path": _secret_text("IPI_GITHUB_PATH"),
        "ref": _secret_text("IPI_GITHUB_REF", "main") or "main",
        "token": _secret_text("IPI_GITHUB_TOKEN"),
    }


def get_label_copy_template_config():
    """
    Reads the private GitHub location of the Label Copy DOCX template.

    Preferred Streamlit secrets:
      LABELCOPY_GITHUB_OWNER
      LABELCOPY_GITHUB_REPO
      LABELCOPY_GITHUB_PATH
      LABELCOPY_GITHUB_TOKEN
      LABELCOPY_GITHUB_REF       -> optional, defaults to "main"

    Each LABELCOPY_* value falls back to the corresponding IPI_GITHUB_* value,
    allowing both private files to live in the same repository. In normal
    deployments ``LABELCOPY_GITHUB_PATH`` should point specifically to the DOCX
    template even when the remaining settings reuse the IPI repository.
    """
    fallback_pairs = {
        "owner": ("LABELCOPY_GITHUB_OWNER", "IPI_GITHUB_OWNER"),
        "repo": ("LABELCOPY_GITHUB_REPO", "IPI_GITHUB_REPO"),
        "path": ("LABELCOPY_GITHUB_PATH", "IPI_GITHUB_PATH"),
        "ref": ("LABELCOPY_GITHUB_REF", "IPI_GITHUB_REF"),
        "token": ("LABELCOPY_GITHUB_TOKEN", "IPI_GITHUB_TOKEN"),
    }

    config: dict[str, str] = {}
    missing: list[str] = []
    for field, (primary_key, fallback_key) in fallback_pairs.items():
        default = "main" if field == "ref" else ""
        value = _secret_text(primary_key) or _secret_text(fallback_key, default)
        if field == "ref":
            value = value or "main"
        elif not value:
            missing.append(primary_key)
        config[field] = value

    if missing:
        raise RuntimeError(
            "Λείπουν Streamlit secrets για το Label Copy template: " + ", ".join(missing)
        )
    return config


# --------------------------------------------------------------------------
# Generic private-file fetcher
# --------------------------------------------------------------------------
def _github_contents_api_url(owner, repo, path):
    encoded_path = "/".join(
        urllib.parse.quote(part, safe="")
        for part in str(path).strip("/").split("/")
        if part
    )
    return f"https://api.github.com/repos/{owner}/{repo}/contents/{encoded_path}"


def _retry_after_seconds(value: Any, default: float = 1.0) -> float:
    try:
        return max(0.0, min(float(value), 30.0))
    except (TypeError, ValueError):
        return default


def _decode_github_contents_response(response: requests.Response) -> bytes:
    content_type = response.headers.get("Content-Type", "").lower()
    if "application/json" not in content_type:
        return response.content

    try:
        data = response.json()
    except ValueError as exc:
        raise PrivateFileFetchError(
            "invalid_response",
            "Το GitHub API επέστρεψε μη έγκυρη JSON απάντηση.",
        ) from exc

    encoded_content = str(data.get("content") or "").replace("\n", "")
    if not encoded_content:
        raise PrivateFileFetchError(
            "empty_content",
            "Το GitHub API δεν επέστρεψε περιεχόμενο για το ζητούμενο αρχείο.",
        )

    try:
        return base64.b64decode(encoded_content, validate=True)
    except (ValueError, TypeError) as exc:
        raise PrivateFileFetchError(
            "invalid_base64",
            "Το GitHub API επέστρεψε μη έγκυρο Base64 περιεχόμενο.",
        ) from exc


@st.cache_data(ttl=PRIVATE_FILE_CACHE_TTL_SECONDS, show_spinner=False)
def fetch_private_file_bytes(
    owner,
    repo,
    path,
    ref,
    token,
    expect_magic=b"PK",
):
    """
    Fetches arbitrary bytes from a private GitHub repository via Contents API.

    Args:
        owner: GitHub user or organization.
        repo: Private repository name.
        path: Repository-relative file path.
        ref: Branch, tag, or commit SHA.
        token: Fine-grained PAT with Contents: Read-only permission.
        expect_magic: Optional byte prefix used as a light file sanity check.
            DOCX and XLSX are ZIP containers and therefore use ``b"PK"``.
            Pass ``None`` or ``b""`` to disable the check.

    Returns:
        The raw file bytes.

    Raises:
        PrivateFileFetchError: Structured, Greek diagnostic suitable for a
        caller-side ``try/except``. The function intentionally mirrors the
        existing IPI fetcher's raising contract because page code already
        handles it explicitly.
    """
    owner = str(owner or "").strip()
    repo = str(repo or "").strip()
    path = str(path or "").strip()
    ref = str(ref or "main").strip() or "main"
    token = str(token or "").strip()

    missing = [
        name
        for name, value in (
            ("owner", owner),
            ("repo", repo),
            ("path", path),
            ("token", token),
        )
        if not value
    ]
    if missing:
        raise PrivateFileFetchError(
            "configuration",
            "Λείπουν στοιχεία GitHub για το ιδιωτικό αρχείο: " + ", ".join(missing),
        )

    url = _github_contents_api_url(owner, repo, path)
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github.raw+json",
        "X-GitHub-Api-Version": GITHUB_API_VERSION,
        "User-Agent": GITHUB_USER_AGENT,
    }

    last_error: PrivateFileFetchError | None = None
    for attempt in range(GITHUB_CONTENTS_MAX_RETRIES):
        try:
            response = requests.get(
                url,
                headers=headers,
                params={"ref": ref},
                timeout=GITHUB_CONTENTS_TIMEOUT_SECONDS,
            )
        except requests.RequestException as exc:
            last_error = PrivateFileFetchError(
                "network",
                f"Αποτυχία επικοινωνίας με το GitHub Contents API: {exc}",
            )
            if attempt + 1 < GITHUB_CONTENTS_MAX_RETRIES:
                time.sleep(0.5 * (2 ** attempt))
                continue
            raise last_error from exc

        if response.status_code in (401, 403):
            raise PrivateFileFetchError(
                "permission",
                "Το GitHub token δεν έχει πρόσβαση στο ιδιωτικό repository.",
            )
        if response.status_code == 404:
            raise PrivateFileFetchError(
                "not_found",
                "Δεν βρέθηκε το ζητούμενο αρχείο στο ιδιωτικό GitHub repository.",
            )
        if response.status_code == 429:
            last_error = PrivateFileFetchError(
                "rate_limit",
                "Το GitHub API επέβαλε προσωρινό περιορισμό κλήσεων.",
            )
            if attempt + 1 < GITHUB_CONTENTS_MAX_RETRIES:
                time.sleep(_retry_after_seconds(response.headers.get("Retry-After")))
                continue
            raise last_error
        if 500 <= response.status_code < 600:
            last_error = PrivateFileFetchError(
                "server",
                f"Το GitHub API επέστρεψε HTTP {response.status_code}.",
            )
            if attempt + 1 < GITHUB_CONTENTS_MAX_RETRIES:
                time.sleep(0.5 * (2 ** attempt))
                continue
            raise last_error
        if not response.ok:
            raise PrivateFileFetchError(
                "http",
                f"Το GitHub API επέστρεψε HTTP {response.status_code}.",
            )

        file_bytes = _decode_github_contents_response(response)
        if not file_bytes:
            raise PrivateFileFetchError(
                "empty_file",
                "Το αρχείο που φορτώθηκε από το GitHub είναι κενό.",
            )

        if expect_magic and not file_bytes.startswith(bytes(expect_magic)):
            magic_display = bytes(expect_magic).decode("ascii", errors="replace")
            raise PrivateFileFetchError(
                "invalid_magic",
                "Το αρχείο που φορτώθηκε από το GitHub δεν έχει την αναμενόμενη "
                f"υπογραφή ({magic_display}).",
            )
        return file_bytes

    if last_error is not None:
        raise last_error
    raise PrivateFileFetchError("unknown", "Αποτυχία φόρτωσης του ιδιωτικού αρχείου.")


def fetch_private_ipi_list_bytes(owner, repo, path, ref, token):
    """
    Backwards-compatible thin wrapper for the private IPI LIST Excel.

    The function retains the original arguments and caller-visible Greek error
    messages so ``tools/page_catalog.py`` remains behaviourally unchanged.
    """
    try:
        return fetch_private_file_bytes(
            owner=owner,
            repo=repo,
            path=path,
            ref=ref,
            token=token,
            expect_magic=b"PK",
        )
    except PrivateFileFetchError as exc:
        if exc.code == "permission":
            raise RuntimeError(
                "Το GitHub token δεν έχει πρόσβαση στο ιδιωτικό IPI LIST repo."
            ) from exc
        if exc.code == "not_found":
            raise RuntimeError(
                "Δεν βρέθηκε το IPI LIST αρχείο στο ιδιωτικό GitHub repo."
            ) from exc
        if exc.code in {"empty_content", "empty_file"}:
            raise RuntimeError("Το IPI LIST αρχείο είναι κενό.") from exc
        if exc.code == "invalid_magic":
            raise RuntimeError(
                "Το αρχείο που φορτώθηκε από GitHub δεν φαίνεται να είναι έγκυρο .xlsx."
            ) from exc
        raise RuntimeError(str(exc)) from exc


def fetch_private_label_copy_template_bytes(owner, repo, path, ref, token):
    """Thin wrapper for the private Label Copy DOCX template."""
    try:
        return fetch_private_file_bytes(
            owner=owner,
            repo=repo,
            path=path,
            ref=ref,
            token=token,
            expect_magic=b"PK",
        )
    except PrivateFileFetchError as exc:
        if exc.code == "permission":
            raise RuntimeError(
                "Το GitHub token δεν έχει πρόσβαση στο ιδιωτικό Label Copy repo."
            ) from exc
        if exc.code == "not_found":
            raise RuntimeError(
                "Δεν βρέθηκε το Label Copy template στο ιδιωτικό GitHub repo."
            ) from exc
        if exc.code in {"empty_content", "empty_file"}:
            raise RuntimeError("Το Label Copy template είναι κενό.") from exc
        if exc.code == "invalid_magic":
            raise RuntimeError(
                "Το αρχείο που φορτώθηκε από GitHub δεν φαίνεται να είναι έγκυρο .docx."
            ) from exc
        raise RuntimeError(str(exc)) from exc


# --------------------------------------------------------------------------
# IPI LIST lookup helpers (existing behaviour retained)
# --------------------------------------------------------------------------
def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _lookup_key(value):
    """Case-insensitive lookup key matching nickname.strip().lower()."""
    text = _clean_text(value)
    text = re.sub(r"\s+", " ", text)
    return text.lower()


def _parse_ipi(value):
    """Return IPI as an int where possible, so Excel stores it as a number."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else int(value)

    text = str(value).strip()
    if not text:
        return None

    compact = text.replace(" ", "").replace(",", "")
    try:
        number = Decimal(compact)
        if number == number.to_integral_value():
            return int(number)
    except (InvalidOperation, ValueError):
        pass

    digits_only = re.sub(r"\D+", "", text)
    if digits_only:
        return int(digits_only)

    return None


@st.cache_data(show_spinner=False)
def build_ipi_lookup_from_bytes(file_bytes):
    """
    Builds:
      {
        nickname.strip().lower(): {"legal": LEGAL NAME, "ipi": IPI, "pro": PRO},
        legal_name.strip().lower(): {"legal": LEGAL NAME, "ipi": IPI, "pro": PRO},
      }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        if "IPI LIST" not in wb.sheetnames:
            raise ValueError("Το αρχείο δεν περιέχει sheet με όνομα 'IPI LIST'.")

        ws = wb["IPI LIST"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("Το sheet 'IPI LIST' δεν έχει headers.")

        header_map = {
            str(header).strip().upper(): idx
            for idx, header in enumerate(header_row)
            if header is not None and str(header).strip()
        }
        required_headers = ["NICKNAME", "LEGAL NAME", "IPI", "PRO"]
        missing_headers = [h for h in required_headers if h not in header_map]
        if missing_headers:
            raise ValueError("Λείπουν headers από το IPI LIST: " + ", ".join(missing_headers))

        def get_cell(row, header_name):
            idx = header_map[header_name]
            return row[idx] if idx < len(row) else None

        lookup = {}
        source_row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            nickname = _clean_text(get_cell(row, "NICKNAME"))
            legal_name = _clean_text(get_cell(row, "LEGAL NAME"))
            ipi = _parse_ipi(get_cell(row, "IPI"))
            pro = _clean_text(get_cell(row, "PRO"))

            if not nickname and not legal_name:
                continue

            entry = {
                "legal": legal_name or nickname,
                "ipi": ipi,
                "pro": pro,
            }

            nickname_key = _lookup_key(nickname)
            legal_key = _lookup_key(legal_name)

            if nickname_key and nickname_key not in lookup:
                lookup[nickname_key] = entry
            if legal_key and legal_key not in lookup:
                lookup[legal_key] = entry

            source_row_count += 1

        return lookup, source_row_count

    finally:
        wb.close()


def _contributor_row(name, ipi_lookup):
    raw_name = _clean_text(name)
    match = ipi_lookup.get(_lookup_key(raw_name)) if ipi_lookup else None
    if match:
        return {
            "raw": raw_name,
            "writer": match.get("legal") or raw_name,
            "ipi": match.get("ipi"),
            "pro": match.get("pro") or "",
            "matched": True,
        }

    return {
        "raw": raw_name,
        "writer": raw_name,
        "ipi": None,
        "pro": "",
        "matched": False,
    }


def _build_contributor_rows(names, ipi_lookup):
    rows = []
    seen = set()

    for name in names:
        if not _clean_text(name):
            continue

        row = _contributor_row(name, ipi_lookup)
        key = _lookup_key(row["writer"])
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)

    return rows


def _scan_ipi_health(ipi_lookup):
    """
    Deduplicates the IPI lookup by legal name and reports entries that are
    missing an IPI number or a PRO affiliation. Returns (total, problems).
    """
    seen = set()
    total = 0
    problems = []
    for entry in ipi_lookup.values():
        legal = (entry.get("legal") or "").strip()
        key = legal.lower()
        if not legal or key in seen:
            continue
        seen.add(key)
        total += 1

        missing = []
        if entry.get("ipi") is None:
            missing.append("IPI")
        if not (entry.get("pro") or "").strip():
            missing.append("PRO")

        if missing:
            problems.append({
                "Writer": legal,
                "IPI": entry.get("ipi") if entry.get("ipi") is not None else "—",
                "PRO": entry.get("pro") or "—",
                "Πρόβλημα": ", ".join(f"Missing {m}" for m in missing),
            })
    return total, problems


__all__ = [
    "PrivateFileFetchError",
    "_build_contributor_rows",
    "_contributor_row",
    "_github_contents_api_url",
    "_lookup_key",
    "_scan_ipi_health",
    "build_ipi_lookup_from_bytes",
    "fetch_private_file_bytes",
    "fetch_private_ipi_list_bytes",
    "fetch_private_label_copy_template_bytes",
    "get_label_copy_template_config",
    "get_private_ipi_config",
]
