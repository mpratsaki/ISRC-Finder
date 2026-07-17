#!/usr/bin/env python3
"""
app.py
Stay Independent Tool - The Stay Independent Catalog Utility
Streamlit web version (Playlist-to-Excel Generator)

Version 2.0 - "Swiss Army Knife Edition"
- Modular, state-based page routing (each page is a standalone function).
- Scalable custom sidebar navigation with active-page highlighting.
- New tools: Musixmatch Sync Checker, Metadata Health, Smart Links (Odesli),
  MusicBrainz Explorer (under development), Settings.
- Backend logic (Spotify auth, Tidal ISRC lookup, private IPI LIST loading,
  Supabase queries, Excel generation) is UNCHANGED.

IMPORTANT (post Feb-2026 Spotify API changes):
Spotify no longer returns playlist contents via Client Credentials, and even
with a logged-in user, playlist items are only returned for playlists that
user OWNS or COLLABORATES ON. So this app makes each visitor log in with
their own Spotify account, and lets them pick from THEIR playlists only.
"""

import base64
import io
import os
import re
import time
import unicodedata
import urllib.parse
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
import pandas as pd
import requests
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from supabase import create_client, Client

SCOPE = "playlist-read-private playlist-read-collaborative"

# unofficial public Tidal client token
TIDAL_TOKEN = "gsFXkJqGrUNoYMQPZe4k3WKwijnrp8iGSwn3bApe"
TIDAL_COUNTRY_CODE = "US"
TIDAL_REQUEST_TIMEOUT_SECONDS = 10
TIDAL_SLEEP_BETWEEN_CALLS_SECONDS = 0.1

TIDAL_ALLOWED_ROLE_KEYS = {"composer", "lyricist", "writer", "author", "producer"}
TIDAL_EXCLUDED_ROLE_KEYS = {
    "musicpublisher",
    "mixingengineer",
    "masteringengineer",
    "recordingengineer",
    "programmer",
    "musician",
    "featured",
    "mainartist",
}

# Private GitHub IPI LIST source. The Excel file must live in a private
# repository. Store the read-only token and file location in Streamlit secrets.
GITHUB_API_VERSION = "2022-11-28"
GITHUB_CONTENTS_TIMEOUT_SECONDS = 20
IPI_LIST_CACHE_TTL_SECONDS = 15 * 60

APP_VERSION = "Stay Independent Tool v2.0 - Swiss Army Knife Edition"


# --------------------------------------------------------------------------
# Credentials & config
# --------------------------------------------------------------------------
def get_config():
    """
    Reads Spotify config from Streamlit secrets (Settings -> Secrets on Streamlit
    Community Cloud). Required keys:
      SPOTIFY_CLIENT_ID
      SPOTIFY_CLIENT_SECRET
      REDIRECT_URI   -> must exactly match a Redirect URI registered in your
                         Spotify Dashboard app, e.g. https://yourapp.streamlit.app
                         (use http://localhost:8501 while testing locally)
    """
    try:
        client_id = st.secrets["SPOTIFY_CLIENT_ID"]
        client_secret = st.secrets["SPOTIFY_CLIENT_SECRET"]
        redirect_uri = st.secrets["REDIRECT_URI"]
    except KeyError as e:
        st.error(f"Λείπει η ρύθμιση {e} από τα Streamlit secrets.")
        st.stop()
    return client_id, client_secret, redirect_uri


def get_private_ipi_config():
    """
    Reads the private GitHub source for the IPI LIST ground-truth Excel.

    Required Streamlit secrets:
      IPI_GITHUB_OWNER  -> GitHub user/org that owns the private repository
      IPI_GITHUB_REPO   -> private repository name
      IPI_GITHUB_PATH   -> path to the xlsx file inside the repository
      IPI_GITHUB_TOKEN  -> fine-grained PAT with Contents: Read-only on that repo

    Optional Streamlit secret:
      IPI_GITHUB_REF    -> branch, tag, or commit SHA. Defaults to "main".
    """
    required_keys = [
        "IPI_GITHUB_OWNER",
        "IPI_GITHUB_REPO",
        "IPI_GITHUB_PATH",
        "IPI_GITHUB_TOKEN",
    ]
    missing = [key for key in required_keys if not str(st.secrets.get(key, "")).strip()]
    if missing:
        raise RuntimeError("Λείπουν Streamlit secrets για το IPI LIST: " + ", ".join(missing))

    return {
        "owner": str(st.secrets["IPI_GITHUB_OWNER"]).strip(),
        "repo": str(st.secrets["IPI_GITHUB_REPO"]).strip(),
        "path": str(st.secrets["IPI_GITHUB_PATH"]).strip(),
        "ref": str(st.secrets.get("IPI_GITHUB_REF", "main")).strip() or "main",
        "token": str(st.secrets["IPI_GITHUB_TOKEN"]).strip(),
    }


# --------------------------------------------------------------------------
# Spotify authentication (Authorization Code flow - per-user login)
# --------------------------------------------------------------------------
def build_authorize_url(client_id, redirect_uri):
    params = {
        "client_id": client_id,
        "response_type": "code",
        "redirect_uri": redirect_uri,
        "scope": SCOPE,
    }
    return "https://accounts.spotify.com/authorize?" + urllib.parse.urlencode(params)


def exchange_code_for_token(client_id, client_secret, redirect_uri, code):
    resp = requests.post(
        "https://accounts.spotify.com/api/token",
        data={
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": redirect_uri,
        },
        auth=(client_id, client_secret),
    )
    resp.raise_for_status()
    data = resp.json()
    data["expires_at"] = time.time() + data.get("expires_in", 3600)
    return data


def refresh_access_token(client_id, client_secret, refresh_token):
    resp = requests.post(
        "https://accounts.spotify.com/api/token",
        data={"grant_type": "refresh_token", "refresh_token": refresh_token},
        auth=(client_id, client_secret),
    )
    resp.raise_for_status()
    data = resp.json()
    data.setdefault("refresh_token", refresh_token)
    data["expires_at"] = time.time() + data.get("expires_in", 3600)
    return data


def get_valid_token():
    """Returns a valid access token from session_state, refreshing if needed."""
    token_data = st.session_state.get("token_data")
    if not token_data:
        return None

    if token_data["expires_at"] > time.time() + 30:
        return token_data["access_token"]

    client_id, client_secret, _ = get_config()
    try:
        token_data = refresh_access_token(client_id, client_secret, token_data["refresh_token"])
        st.session_state["token_data"] = token_data
        return token_data["access_token"]
    except requests.HTTPError:
        st.session_state.pop("token_data", None)
        return None


# --------------------------------------------------------------------------
# Spotify data fetching & validation
# --------------------------------------------------------------------------
def extract_playlist_id(playlist_arg):
    m = re.search(r"playlist[/:]([a-zA-Z0-9]+)", playlist_arg)
    if m:
        return m.group(1).split("?")[0]
    return playlist_arg.strip()


def _api_get(token, url, params=None, retries=3):
    headers = {"Authorization": f"Bearer {token}"}
    for attempt in range(retries):
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code == 429:
            wait = int(resp.headers.get("Retry-After", 2))
            time.sleep(wait)
            continue
        resp.raise_for_status()
        return resp.json()
    resp.raise_for_status()


def fetch_user_playlists(token):
    """Returns the logged-in user's own + collaborative playlists."""
    playlists = []
    url = "https://api.spotify.com/v1/me/playlists"
    params = {"limit": 50}

    while url:
        data = _api_get(token, url, params=params)
        for item in data.get("items", []):
            playlists.append({"id": item["id"], "name": item["name"]})
        url = data.get("next")
        params = None

    return playlists


def fetch_playlist_tracks(token, playlist_id):
    tracks = []
    url = f"https://api.spotify.com/v1/playlists/{playlist_id}/items"
    params = {
        "fields": "items(item(id,name,artists(name),external_ids)),next",
        "limit": 50,
        "offset": 0,
    }

    while url:
        data = _api_get(token, url, params=params)
        items = data.get("items")
        if items is None:
            raise RuntimeError(
                "Δεν επιστράφηκε περιεχόμενο playlist. Βεβαιωθείτε ότι η playlist "
                "είναι δική σας ή collaborative."
            )

        for entry in items:
            track = entry.get("item")
            if not track:
                continue
            isrc = (track.get("external_ids") or {}).get("isrc")
            tracks.append(
                {
                    "id": track["id"],
                    "name": track["name"],
                    "artists": [a["name"] for a in track.get("artists", [])],
                    "isrc": isrc,
                }
            )

        url = data.get("next")
        params = None

    return tracks


def validate_isrc(isrc):
    if not isrc:
        return False
    clean_isrc = str(isrc).replace("-", "").strip()
    pattern = re.compile(r"^[A-Z]{2}[A-Z0-9]{3}\d{2}\d{5}$", re.IGNORECASE)
    return bool(pattern.match(clean_isrc))


# --------------------------------------------------------------------------
# Tidal credits lookup by ISRC
# --------------------------------------------------------------------------
def _role_key(role):
    return re.sub(r"[\s_\-]+", "", str(role or "").strip()).lower()


def _is_allowed_tidal_role(role):
    """Allow only songwriter/producer roles; exclude publishers, engineers, artists, etc."""
    role_text = str(role or "").strip()
    if not role_text:
        return False

    key = _role_key(role_text)
    if key in TIDAL_EXCLUDED_ROLE_KEYS:
        return False
    if key in TIDAL_ALLOWED_ROLE_KEYS:
        return True

    # Defensive handling for rare combined role strings such as "Composer/Lyricist".
    parts = re.split(r"[,;/|&]+", role_text)
    return any(_role_key(part) in TIDAL_ALLOWED_ROLE_KEYS for part in parts)


def _extract_items(data):
    """Tidal responses normally contain an 'items' list; keep this tolerant."""
    if isinstance(data, list):
        return data
    if not isinstance(data, dict):
        return []

    items = data.get("items")
    if isinstance(items, list):
        return items

    tracks = data.get("tracks")
    if isinstance(tracks, dict) and isinstance(tracks.get("items"), list):
        return tracks["items"]

    return []


def _tidal_get(url, params=None, retries=3):
    """
    Safe Tidal GET wrapper. Never raises to the UI path.
    Returns (json_data, note_if_failed).
    """
    headers = {"X-Tidal-Token": TIDAL_TOKEN}
    last_note = "Tidal request failed — used Spotify artists as fallback"

    for _ in range(retries):
        try:
            resp = requests.get(
                url,
                headers=headers,
                params=params,
                timeout=TIDAL_REQUEST_TIMEOUT_SECONDS,
            )

            if resp.status_code == 429:
                retry_after = resp.headers.get("Retry-After", "1")
                try:
                    wait_seconds = max(float(retry_after), 0.0)
                except (TypeError, ValueError):
                    wait_seconds = 1.0
                time.sleep(wait_seconds)
                last_note = "Tidal rate limit — used Spotify artists as fallback"
                continue

            if not resp.ok:
                return None, f"Tidal HTTP {resp.status_code} — used Spotify artists as fallback"

            return resp.json(), None

        except requests.RequestException:
            last_note = "Tidal request failed — used Spotify artists as fallback"
        except ValueError:
            return None, "Tidal response invalid — used Spotify artists as fallback"
        finally:
            time.sleep(TIDAL_SLEEP_BETWEEN_CALLS_SECONDS)

    return None, last_note


def fetch_tidal_contributors_by_isrc(isrc):
    """
    Finds the first Tidal track by ISRC and returns unique contributor names for
    Composer/Lyricist/Writer/Author/Producer roles only.
    Returns (names, note_if_fallback_needed).
    """
    try:
        clean_isrc = str(isrc or "").replace("-", "").strip().upper()
        if not clean_isrc:
            return [], "Missing ISRC — used Spotify artists as fallback"
        if not validate_isrc(clean_isrc):
            return [], "ISRC format invalid — used Spotify artists as fallback"

        search_data, note = _tidal_get(
            "https://api.tidal.com/v1/tracks",
            params={"isrc": clean_isrc, "countryCode": TIDAL_COUNTRY_CODE},
        )
        if note or not search_data:
            return [], note or "Tidal credits not found — used Spotify artists as fallback"

        track_items = _extract_items(search_data)
        if not track_items:
            return [], "Tidal credits not found — used Spotify artists as fallback"

        tidal_track_id = track_items[0].get("id") if isinstance(track_items[0], dict) else None
        if not tidal_track_id:
            return [], "Tidal track ID missing — used Spotify artists as fallback"

        contributors_data, note = _tidal_get(
            f"https://api.tidal.com/v1/tracks/{tidal_track_id}/contributors",
            params={"countryCode": TIDAL_COUNTRY_CODE},
        )
        if note or not contributors_data:
            return [], note or "Tidal credits not found — used Spotify artists as fallback"

        contributor_items = _extract_items(contributors_data)
        if not contributor_items:
            return [], "Tidal credits not found — used Spotify artists as fallback"

        names = []
        seen = set()
        for item in contributor_items:
            if not isinstance(item, dict):
                continue

            name = str(item.get("name") or "").strip()
            role = item.get("role")
            if not name or not _is_allowed_tidal_role(role):
                continue

            key = _lookup_key(name)
            if key in seen:
                continue
            seen.add(key)
            names.append(name)

        if not names:
            return [], "Tidal credits not found — used Spotify artists as fallback"

        return names, None

    except Exception:
        return [], "Tidal lookup failed — used Spotify artists as fallback"


# --------------------------------------------------------------------------
# Private GitHub IPI LIST source
# --------------------------------------------------------------------------
def _github_contents_api_url(owner, repo, path):
    encoded_path = "/".join(
        urllib.parse.quote(part, safe="")
        for part in str(path).strip("/").split("/")
        if part
    )
    return f"https://api.github.com/repos/{owner}/{repo}/contents/{encoded_path}"


@st.cache_data(ttl=IPI_LIST_CACHE_TTL_SECONDS, show_spinner=False)
def fetch_private_ipi_list_bytes(owner, repo, path, ref, token):
    """
    Fetches the private IPI LIST Excel from GitHub using the repository contents
    API. The token must be stored in Streamlit secrets, not in the repository.
    """
    url = _github_contents_api_url(owner, repo, path)
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github.raw+json",
        "X-GitHub-Api-Version": GITHUB_API_VERSION,
        "User-Agent": "stay-independent-catalog-generator",
    }

    response = requests.get(
        url,
        headers=headers,
        params={"ref": ref},
        timeout=GITHUB_CONTENTS_TIMEOUT_SECONDS,
    )

    if response.status_code in (401, 403):
        raise RuntimeError("Το GitHub token δεν έχει πρόσβαση στο ιδιωτικό IPI LIST repo.")
    if response.status_code == 404:
        raise RuntimeError("Δεν βρέθηκε το IPI LIST αρχείο στο ιδιωτικό GitHub repo.")

    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "").lower()
    if "application/json" in content_type:
        # Defensive fallback if GitHub returns the default JSON representation
        # instead of raw bytes. The 'content' field is Base64 encoded.
        data = response.json()
        encoded_content = str(data.get("content") or "").replace("\n", "")
        if not encoded_content:
            raise RuntimeError("Το GitHub API δεν επέστρεψε περιεχόμενο για το IPI LIST.")
        file_bytes = base64.b64decode(encoded_content)
    else:
        file_bytes = response.content

    if not file_bytes:
        raise RuntimeError("Το IPI LIST αρχείο είναι κενό.")

    # .xlsx files are ZIP containers and normally start with PK. This catches
    # accidental HTML/JSON error payloads before openpyxl tries to parse them.
    if not file_bytes.startswith(b"PK"):
        raise RuntimeError("Το αρχείο που φορτώθηκε από GitHub δεν φαίνεται να είναι έγκυρο .xlsx.")

    return file_bytes


# --------------------------------------------------------------------------
# IPI LIST lookup helpers
# --------------------------------------------------------------------------
def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _lookup_key(value):
    """Case-insensitive lookup key matching nickname.strip().lower()."""
    text = _clean_text(value)
    text = re.sub(r"\s+", " ", text)
    return text.lower()


def _parse_ipi(value):
    """Return IPI as an int where possible, so Excel stores it as a number."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else int(value)

    text = str(value).strip()
    if not text:
        return None

    compact = text.replace(" ", "").replace(",", "")
    try:
        number = Decimal(compact)
        if number == number.to_integral_value():
            return int(number)
    except (InvalidOperation, ValueError):
        pass

    digits_only = re.sub(r"\D+", "", text)
    if digits_only:
        return int(digits_only)

    return None


@st.cache_data(show_spinner=False)
def build_ipi_lookup_from_bytes(file_bytes):
    """
    Builds:
      {
        nickname.strip().lower(): {"legal": LEGAL NAME, "ipi": IPI, "pro": PRO},
        legal_name.strip().lower(): {"legal": LEGAL NAME, "ipi": IPI, "pro": PRO},
      }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        if "IPI LIST" not in wb.sheetnames:
            raise ValueError("Το αρχείο δεν περιέχει sheet με όνομα 'IPI LIST'.")

        ws = wb["IPI LIST"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("Το sheet 'IPI LIST' δεν έχει headers.")

        header_map = {
            str(header).strip().upper(): idx
            for idx, header in enumerate(header_row)
            if header is not None and str(header).strip()
        }
        required_headers = ["NICKNAME", "LEGAL NAME", "IPI", "PRO"]
        missing_headers = [h for h in required_headers if h not in header_map]
        if missing_headers:
            raise ValueError("Λείπουν headers από το IPI LIST: " + ", ".join(missing_headers))

        def get_cell(row, header_name):
            idx = header_map[header_name]
            return row[idx] if idx < len(row) else None

        lookup = {}
        source_row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            nickname = _clean_text(get_cell(row, "NICKNAME"))
            legal_name = _clean_text(get_cell(row, "LEGAL NAME"))
            ipi = _parse_ipi(get_cell(row, "IPI"))
            pro = _clean_text(get_cell(row, "PRO"))

            if not nickname and not legal_name:
                continue

            entry = {
                "legal": legal_name or nickname,
                "ipi": ipi,
                "pro": pro,
            }

            nickname_key = _lookup_key(nickname)
            legal_key = _lookup_key(legal_name)

            if nickname_key and nickname_key not in lookup:
                lookup[nickname_key] = entry
            if legal_key and legal_key not in lookup:
                lookup[legal_key] = entry

            source_row_count += 1

        return lookup, source_row_count

    finally:
        wb.close()


def _contributor_row(name, ipi_lookup):
    raw_name = _clean_text(name)
    match = ipi_lookup.get(_lookup_key(raw_name)) if ipi_lookup else None
    if match:
        return {
            "raw": raw_name,
            "writer": match.get("legal") or raw_name,
            "ipi": match.get("ipi"),
            "pro": match.get("pro") or "",
            "matched": True,
        }

    return {
        "raw": raw_name,
        "writer": raw_name,
        "ipi": None,
        "pro": "",
        "matched": False,
    }


def _build_contributor_rows(names, ipi_lookup):
    rows = []
    seen = set()

    for name in names:
        if not _clean_text(name):
            continue

        row = _contributor_row(name, ipi_lookup)
        key = _lookup_key(row["writer"])
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)

    return rows


# --------------------------------------------------------------------------
# Excel Generator (Zero-to-Excel) - builds in-memory
# --------------------------------------------------------------------------
def generate_new_catalog(tracks, ipi_lookup=None, progress_callback=None):
    ipi_lookup = ipi_lookup or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stay Independent Catalog"

    report = {
        "filled": [],
        "health_warnings": [],
        "tidal_fallbacks": [],
        "ipi_matches": 0,
    }

    # --- Ορισμός Στυλ (Γραμματοσειρές, Στοιχίσεις, Χρώματα, Περιγράμματα) ---
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    top_alignment = Alignment(vertical="top", wrap_text=True)
    sum_font = Font(bold=True, color="000000")  # Έντονη γραφή για το άθροισμα

    # Πιο έντονο μαύρο περίγραμμα (medium style)
    black_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )

    # Γκρι γέμισμα για το διαχωριστικό κελί
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Νέα σειρά στηλών βάσει του πρότυπου Excel
    headers = ["TITLE", "ROLE", "WRITERS", "IPI", "PRO", "% RIGHTS", "ISRC", "NOTES"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = black_border

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 35  # TITLE
    ws.column_dimensions["B"].width = 15  # ROLE
    ws.column_dimensions["C"].width = 32  # WRITERS
    ws.column_dimensions["D"].width = 16  # IPI
    ws.column_dimensions["E"].width = 18  # PRO
    ws.column_dimensions["F"].width = 15  # % RIGHTS
    ws.column_dimensions["G"].width = 20  # ISRC
    ws.column_dimensions["H"].width = 55  # NOTES

    insert_at = 2
    total_tracks = len(tracks)

    for track_index, track in enumerate(tracks, start=1):
        title = track.get("name") or ""
        spotify_artists = track.get("artists") or []
        isrc = track.get("isrc")
        clean_isrc = str(isrc or "").replace("-", "").strip().upper()

        if progress_callback:
            progress_callback(track_index, total_tracks, title)

        notes = []
        tidal_names = []
        tidal_note = None
        should_try_tidal = bool(clean_isrc) and validate_isrc(clean_isrc)

        if clean_isrc and not validate_isrc(clean_isrc):
            report["health_warnings"].append((title, isrc))
            notes.append("ISRC format invalid")
        elif not clean_isrc:
            notes.append("Missing ISRC")

        if should_try_tidal:
            tidal_names, tidal_note = fetch_tidal_contributors_by_isrc(clean_isrc)

        if tidal_names:
            contributor_names = tidal_names
            contributor_source = "tidal"
        else:
            contributor_names = spotify_artists
            contributor_source = "spotify_fallback"

            if should_try_tidal:
                note = tidal_note or "Tidal credits not found — used Spotify artists as fallback"
                notes.append(note)
                report["tidal_fallbacks"].append(title)
            elif clean_isrc:
                notes.append("Tidal lookup skipped — used Spotify artists as fallback")
            else:
                notes.append("Tidal lookup skipped — used Spotify artists as fallback")

        contributor_rows = _build_contributor_rows(contributor_names, ipi_lookup)
        report["ipi_matches"] += sum(1 for row in contributor_rows if row["matched"])

        needed_rows = max(1, len(contributor_rows))
        notes_text = "; ".join(dict.fromkeys(note for note in notes if note))

        # Αποθηκεύουμε την αρχική και τελική γραμμή του τρέχοντος τραγουδιού για τη συνάρτηση SUM
        start_row = insert_at
        end_row = insert_at + needed_rows - 1

        for i in range(needed_rows):
            current_row = insert_at + i

            # Επαναλαμβάνουμε TITLE, ISRC και NOTES σε ΚΑΘΕ γραμμή του πλαισίου
            ws.cell(row=current_row, column=1).value = title

            if isrc:
                ws.cell(row=current_row, column=7).value = isrc  # Το ISRC πήγε στη στήλη 7 (G)
            if notes_text:
                ws.cell(row=current_row, column=8).value = notes_text  # Τα NOTES πήγαν στη στήλη 8 (H)

            if i < len(contributor_rows):
                contributor = contributor_rows[i]
                ws.cell(row=current_row, column=3).value = contributor["writer"]  # WRITERS στήλη 3 (C)

                if contributor["ipi"] is not None:
                    ipi_cell = ws.cell(row=current_row, column=4)  # IPI στήλη 4 (D)
                    ipi_cell.value = contributor["ipi"]
                    ipi_cell.number_format = "0"

                if contributor["pro"]:
                    ws.cell(row=current_row, column=5).value = contributor["pro"]  # PRO στήλη 5 (E)

            # Εφαρμογή του μαύρου περιγράμματος σε όλα τα 8 κελιά της τρέχουσας γραμμής
            for col_num in range(1, 9):
                cell = ws.cell(row=current_row, column=col_num)
                cell.alignment = top_alignment
                cell.border = black_border

        # --- Προσθήκη Γκρι Διαχωριστικής Γραμμής & Δυναμικού Υπολογισμού ---
        separator_row = insert_at + needed_rows
        for col_num in range(1, 9):
            cell = ws.cell(row=separator_row, column=col_num)
            cell.fill = gray_fill
            cell.border = black_border

            # Εισαγωγή της δυναμικής Excel Formula (=SUM) στο κελί της στήλης 6 (% RIGHTS)
            if col_num == 6:
                cell.value = f"=SUM(F{start_row}:F{end_row})"
                cell.font = sum_font
                cell.alignment = center_alignment

        report["filled"].append(
            {
                "title": title,
                "contributors": [row["writer"] for row in contributor_rows] if contributor_rows else [],
                "isrc": isrc,
                "source": contributor_source,
                "notes": notes_text,
            }
        )

        # Το επόμενο τραγούδι ξεκινά μετά το block των writers + 1 γραμμή για το διαχωριστικό
        insert_at += needed_rows + 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, report


# --------------------------------------------------------------------------
# Filename helper
# --------------------------------------------------------------------------
def make_catalog_filename(playlist_name):
    safe_name = re.sub(r"\s+", "_", str(playlist_name or "playlist").strip())
    safe_name = unicodedata.normalize("NFKD", safe_name).encode("ascii", "ignore").decode("ascii")
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "", safe_name).strip("_")
    if not safe_name:
        safe_name = "playlist"

    date_part = datetime.now().strftime("%Y%m%d")
    return f"Catalog_{safe_name}_{date_part}.xlsx"


# ==========================================================================
# ==========================================================================
#  STREAMLIT UI LAYER (Swiss Army Knife Edition)
#  Backend above is untouched. Everything below is presentation / routing.
# ==========================================================================
# ==========================================================================

st.set_page_config(
    page_title="Stay Independent Tool",
    page_icon="StayLogo2.jpg",
    layout="wide",
    initial_sidebar_state="expanded",
)

# --- Ενσωμάτωση Custom CSS ---
st.markdown("""
    <style>
    .stButton>button {
        border-radius: 8px;
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }
    .live-activity-box {
        padding: 15px;
        border-radius: 10px;
        background: linear-gradient(135deg, #2a2a2a, #1f1f1f);
        border-left: 5px solid #1DB954;
        margin-bottom: 20px;
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }
    .block-container {
        padding-top: 2rem;
    }
    .under-construction {
        padding: 40px 25px;
        border-radius: 14px;
        text-align: center;
        background: repeating-linear-gradient(
            45deg, #2b2b2b, #2b2b2b 18px, #242424 18px, #242424 36px
        );
        border: 2px dashed #FFB020;
        box-shadow: 0 6px 18px rgba(0,0,0,0.25);
        margin-bottom: 25px;
    }
    .under-construction h2 { color:#FFB020; margin-bottom:8px; }
    .under-construction p  { color:#ddd; font-size:16px; }
    /* Try to give the sidebar a column layout so the system block can sit lower */
    section[data-testid="stSidebar"] div[data-testid="stSidebarUserContent"] {
        display: flex;
        flex-direction: column;
    }
    </style>
""", unsafe_allow_html=True)


# --------------------------------------------------------------------------
# Shared resources / helpers
# --------------------------------------------------------------------------
@st.cache_resource
def init_supabase() -> Client:
    url = st.secrets.get("SUPABASE_URL")
    key = st.secrets.get("SUPABASE_KEY")
    if url and key:
        return create_client(url, key)
    return None


def fetch_current_user(token):
    data = _api_get(token, "https://api.spotify.com/v1/me")
    return data.get("display_name") or data.get("id") or "Άγνωστος Χρήστης"


@st.cache_data(ttl=300, show_spinner=False)
def fetch_odesli_links(track_url):
    """
    Calls the free Odesli (song.link) API and returns the parsed JSON.
    GET https://api.odesli.co/v1-alpha.1/links?url=<encoded url>
    """
    resp = requests.get(
        "https://api.odesli.co/v1-alpha.1/links",
        params={"url": track_url},
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()


def _scan_ipi_health(ipi_lookup):
    """
    Deduplicates the IPI lookup by legal name and reports entries that are
    missing an IPI number or a PRO affiliation. Returns (total, problems).
    """
    seen = set()
    total = 0
    problems = []
    for entry in ipi_lookup.values():
        legal = (entry.get("legal") or "").strip()
        key = legal.lower()
        if not legal or key in seen:
            continue
        seen.add(key)
        total += 1

        missing = []
        if entry.get("ipi") is None:
            missing.append("IPI")
        if not (entry.get("pro") or "").strip():
            missing.append("PRO")

        if missing:
            problems.append({
                "Writer": legal,
                "IPI": entry.get("ipi") if entry.get("ipi") is not None else "—",
                "PRO": entry.get("pro") or "—",
                "Πρόβλημα": ", ".join(f"Missing {m}" for m in missing),
            })
    return total, problems


# ==========================================================================
# PAGE: Landing / Login (no sidebar, centered) - logged-out state
# ==========================================================================
def render_landing_page(client_id, redirect_uri):
    st.markdown("<br><br>", unsafe_allow_html=True)

    # Smaller, aesthetically balanced logo via a narrow center column.
    logo_left, logo_center, logo_right = st.columns([4, 2, 4])
    with logo_center:
        if os.path.exists("StayLogo2.jpg"):
            st.image("StayLogo2.jpg", width="stretch")

    st.markdown(
        "<h1 style='text-align:center;'>Stay Independent Tool</h1>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p style='text-align:center; font-size:18px; color:#aaa;'>"
        "Welcome to the Stay Independent multi-tool. Please log in with your "
        "Spotify account to access the Catalog Generator and your Export History."
        "</p>",
        unsafe_allow_html=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)

    auth_url = build_authorize_url(client_id, redirect_uri)
    _, btn_col, _ = st.columns([1, 2, 1])
    with btn_col:
        st.link_button("Σύνδεση με Spotify", auth_url, type="primary", width="stretch")


# ==========================================================================
# PAGE: Γεννήτρια Catalog (main feature) - unchanged logic
# ==========================================================================
def page_catalog_generator(token, spotify_user):
    st.title("Γεννήτρια Catalog")

    # Φόρτωση IPI List
    try:
        private_ipi_config = get_private_ipi_config()
        with st.spinner("Φόρτωση IPI LIST από το GitHub..."):
            ipi_file_bytes = fetch_private_ipi_list_bytes(**private_ipi_config)
            ipi_lookup, ipi_source_rows = build_ipi_lookup_from_bytes(ipi_file_bytes)
    except Exception as e:
        st.error("Αδυναμία φόρτωσης IPI LIST από το ιδιωτικό repository.")
        st.caption(f"Λεπτομέρεια συστήματος: {e}")
        st.stop()

    # Ανάκτηση Playlists
    try:
        playlists = fetch_user_playlists(token)
    except requests.HTTPError as e:
        st.error(f"Σφάλμα επικοινωνίας με το Spotify: {e}")
        st.stop()

    if not playlists:
        st.warning("Δεν βρέθηκαν playlists στον λογαριασμό σας.")
        st.stop()

    st.markdown("### Επιλογή Δεδομένων")
    playlist_names = [p["name"] for p in playlists]

    col_sel, col_btn = st.columns([3, 1], vertical_alignment="bottom")
    with col_sel:
        selected_name = st.selectbox("Επιλέξτε Playlist για εξαγωγή:", playlist_names)
        selected_playlist = next(p for p in playlists if p["name"] == selected_name)

    with col_btn:
        generate_trigger = st.button("Δημιουργία Catalog", type="primary", width="stretch")

    if generate_trigger:
        st.divider()
        try:
            tracks = fetch_playlist_tracks(token, selected_playlist["id"])

            if not tracks:
                st.warning("Η playlist είναι κενή.")
                st.stop()

            st.markdown("#### Live Activity")
            live_status = st.empty()
            progress_bar = st.progress(0.0)

            def update_generation_progress(current, total, title):
                progress_value = current / max(total, 1)
                progress_bar.progress(progress_value)
                html_content = f"""
                <div class="live-activity-box">
                    <span style="color:#aaa; font-size:14px;">Επεξεργασία {current} από {total}</span><br>
                    <strong style="font-size:18px;">🎵 {title}</strong>
                </div>
                """
                live_status.markdown(html_content, unsafe_allow_html=True)

            buffer, report = generate_new_catalog(
                tracks,
                ipi_lookup=ipi_lookup,
                progress_callback=update_generation_progress,
            )

            live_status.empty()
            progress_bar.empty()
            st.toast("Η δημιουργία του Excel ολοκληρώθηκε!", icon="🎉")

            output_filename = make_catalog_filename(selected_playlist["name"])

            # ΠΡΟΣΘΗΚΗ 2: Αποθήκευση στο Supabase Storage
            supabase = init_supabase()
            file_public_url = None

            if supabase and spotify_user:
                try:
                    # Μετατροπή του BytesIO σε raw bytes
                    file_bytes = buffer.getvalue()

                    # Δημιουργία μοναδικού ονόματος για το αρχείο στο bucket
                    timestamp = int(time.time())
                    storage_path = f"{spotify_user}/{timestamp}_{output_filename}"

                    # Upload στο bucket με όνομα "catalogs" (πρέπει να το φτιάξεις στο Supabase!)
                    supabase.storage.from_("catalogs").upload(
                        file=file_bytes,
                        path=storage_path,
                        file_options={"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
                    )
                    # Ανάκτηση του Public URL
                    file_public_url = supabase.storage.from_("catalogs").get_public_url(storage_path)

                    # Εγγραφή ιστορικού στη Βάση Δεδομένων με το URL
                    supabase.table("export_history").insert({
                        "spotify_user": spotify_user,
                        "playlist_name": selected_playlist["name"],
                        "track_count": len(tracks),
                        "file_url": file_public_url  # Η νέα στήλη
                    }).execute()

                except Exception as e:
                    st.error(f"🚨 Σφάλμα επικοινωνίας με το Supabase: {e}")
                    st.toast("Δεν ενημερώθηκε το ιστορικό.", icon="⚠️")

            # Εμφάνιση Αποτελεσμάτων
            st.markdown("### 📊 Αποτελέσματα & Εξαγωγή")
            tab_summary, tab_preview, tab_logs = st.tabs(["Σύνοψη", "Προεπισκόπηση", "Σφάλματα & Logs"])

            with tab_summary:
                m1, m2, m3 = st.columns(3)
                m1.metric("Σύνολο Τραγουδιών", len(tracks))
                m2.metric("Επιτυχείς Αντιστοιχίσεις IPI", report.get('ipi_matches', 0))
                m3.metric("Προειδοποιήσεις", len(report['health_warnings']) + len(report['tidal_fallbacks']))

                st.markdown("<br>", unsafe_allow_html=True)

                _, col_down, _ = st.columns([1, 2, 1])
                with col_down:
                    st.download_button(
                        label="⬇️ Λήψη Ολοκληρωμένου Excel",
                        data=buffer.getvalue(),
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width="stretch",
                        type="primary"
                    )

            with tab_preview:
                if report["filled"]:
                    df_preview = pd.DataFrame(report["filled"])
                    df_preview["contributors"] = df_preview["contributors"].apply(lambda x: ", ".join(x))
                    st.dataframe(
                        df_preview,
                        width="stretch",
                        hide_index=True,
                        column_config={
                            "title": "Τίτλος",
                            "contributors": "Δημιουργοί",
                            "isrc": "ISRC",
                            "source": "Πηγή Credits",
                            "notes": "Σημειώσεις"
                        }
                    )

            with tab_logs:
                if report["health_warnings"]:
                    st.error(f"Βρέθηκαν {len(report['health_warnings'])} προβληματικά ISRC")
                    for title, isrc in report["health_warnings"]:
                        st.write(f"• **{title}** | ISRC: `{isrc}`")
                else:
                    st.success("Κανένα πρόβλημα με τα ISRC!")

                st.divider()

                if report["tidal_fallbacks"]:
                    st.warning(f"Βρέθηκαν {len(report['tidal_fallbacks'])} τραγούδια χωρίς Tidal Credits")
                    with st.expander("Προβολή λίστας", expanded=False):
                        for title in report["tidal_fallbacks"]:
                            st.write(f"• **{title}**")

        except Exception as e:
            st.error(f"Μη αναμενόμενο σφάλμα κατά τη δημιουργία: {e}")


# ==========================================================================
# PAGE: Ιστορικό & Αρχεία - unchanged logic
# ==========================================================================
def page_history(token, spotify_user):
    st.title("📂 Ιστορικό Εξαγωγών")
    st.markdown("Εδώ μπορείτε να δείτε τις προηγούμενες εξαγωγές σας, να κατεβάσετε τα αρχεία Excel, ή να τα διαγράψετε οριστικά.")

    supabase = init_supabase()
    if not supabase:
        st.warning("Το ιστορικό δεν είναι διαθέσιμο (Λείπουν τα credentials του Supabase στα Secrets).")
        return

    try:
        # Φέρνουμε τα δεδομένα - φέρνουμε και το μοναδικό "id" της εγγραφής
        response = supabase.table("export_history") \
            .select("id, playlist_name, track_count, exported_at, file_url") \
            .eq("spotify_user", spotify_user) \
            .order("exported_at", desc=True) \
            .execute()

        if response.data:
            df_history = pd.DataFrame(response.data)
            df_history["exported_at"] = pd.to_datetime(df_history["exported_at"]).dt.tz_convert("Europe/Athens").dt.strftime("%d-%m-%Y %H:%M:%S")

            # Προσθέτουμε μια προσωρινή στήλη (Checkbox) στην αρχή του Dataframe
            df_history.insert(0, "Επιλογή", False)

            st.markdown("### Λίστα Αρχείων")
            st.caption("Επιλέξτε το κουτάκι αριστερά από τις εγγραφές που θέλετε να διαγράψετε.")

            # Χρήση data_editor για διαδραστικότητα
            edited_df = st.data_editor(
                df_history,
                width="stretch",
                hide_index=True,
                column_config={
                    "id": None,  # Κρύβουμε το ID από τον χρήστη για να είναι καθαρό το UI
                    "Επιλογή": st.column_config.CheckboxColumn("Διαγραφή;", default=False),
                    "playlist_name": st.column_config.TextColumn("Τίτλος Playlist", disabled=True),
                    "track_count": st.column_config.NumberColumn("Τραγούδια", disabled=True),
                    "exported_at": st.column_config.TextColumn("Ημερομηνία & Ώρα", disabled=True),
                    "file_url": st.column_config.LinkColumn(
                        "Αρχείο Excel",
                        help="Πατήστε για να κατεβάσετε το αρχείο",
                        display_text="Κατέβασμα 📥",
                        disabled=True
                    )
                }
            )

            # Εντοπισμός των γραμμών που ο χρήστης τσέκαρε
            rows_to_delete = edited_df[edited_df["Επιλογή"] == True]

            if not rows_to_delete.empty:
                st.warning(f"Έχετε επιλέξει {len(rows_to_delete)} αρχεία προς διαγραφή. Η ενέργεια δεν αναιρείται.")

                if st.button("🗑️ Οριστική Διαγραφή Επιλεγμένων", type="primary"):
                    with st.spinner("Διαγραφή σε εξέλιξη..."):
                        success_count = 0

                        for _, row in rows_to_delete.iterrows():
                            record_id = row["id"]
                            file_url = row["file_url"]

                            # Βήμα 1: Διαγραφή του αρχείου από το Supabase Storage
                            if file_url:
                                try:
                                    # Το URL είναι της μορφής: .../public/catalogs/User/1234_File.xlsx
                                    # Το κόβουμε για να πάρουμε μόνο το "User/1234_File.xlsx"
                                    if "/public/catalogs/" in file_url:
                                        storage_path = file_url.split("/public/catalogs/")[-1]
                                        # Προσοχή: Η remove δέχεται λίστα με paths
                                        supabase.storage.from_("catalogs").remove([storage_path])
                                except Exception as e:
                                    st.error(f"Αδυναμία διαγραφής αρχείου από το Storage: {e}")

                            # Βήμα 2: Διαγραφή της εγγραφής από τη βάση δεδομένων (Table)
                            try:
                                supabase.table("export_history").delete().eq("id", record_id).execute()
                                success_count += 1
                            except Exception as e:
                                st.error(f"Αδυναμία διαγραφής εγγραφής {record_id} από τη βάση: {e}")

                        if success_count > 0:
                            st.success(f"Διαγράφηκαν επιτυχώς {success_count} εγγραφές/αρχεία!")
                            time.sleep(1.5)  # Μικρή παύση για να προλάβει ο χρήστης να διαβάσει το μήνυμα
                            st.rerun()  # Ανανέωση της σελίδας για να ενημερωθεί ο πίνακας
        else:
            st.info("Δεν υπάρχει ιστορικό εξαγωγών για τον λογαριασμό σας.")
    except Exception as e:
        st.error(f"Αδυναμία ανάκτησης ιστορικού: {e}")


# ==========================================================================
# TOOL A: Musixmatch Sync Checker (placeholder UI)
# ==========================================================================
def page_musixmatch_checker():
    st.title("Musixmatch Sync Checker")
    st.caption(
        "Έλεγχος διαθεσιμότητας synced / unsynced lyrics ανά ISRC. "
        "**Demo UI** — τα δεδομένα είναι placeholder μέχρι να συνδεθεί το Musixmatch API."
    )

    mode = st.radio("Λειτουργία εισαγωγής", ["Μονό ISRC", "Πολλαπλά ISRCs"], horizontal=True)

    isrcs = []
    if mode == "Μονό ISRC":
        single = st.text_input("ISRC", placeholder="π.χ. USRC17607839")
        if single.strip():
            isrcs = [single.strip().upper()]
    else:
        multi = st.text_area(
            "ISRCs (ένα ανά γραμμή)",
            height=160,
            placeholder="USRC17607839\nGBUM71029604\nGRA123456789",
        )
        isrcs = [line.strip().upper() for line in multi.splitlines() if line.strip()]

    if st.button("Έλεγχος", type="primary", width="stretch"):
        if not isrcs:
            st.warning("Εισάγετε τουλάχιστον ένα ISRC για έλεγχο.")
            return

        total = len(isrcs)

        # --- Placeholder split (dummy) ---
        synced = round(total * 0.62)
        missing = round(total * 0.15)
        unsynced = max(total - synced - missing, 0)

        st.divider()
        st.markdown("Αναφορά Συγχρονισμού (Demo)")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Σύνολο ISRCs", total)
        c2.metric("Synced Lyrics", synced, delta="OK", delta_color="normal")
        c3.metric("Unsynced", unsynced, delta="-λείπει timing", delta_color="inverse")
        c4.metric("Missing", missing, delta="-χωρίς lyrics", delta_color="inverse")

        st.markdown("<br>", unsafe_allow_html=True)

        st.success(f"✅ {synced} κομμάτια έχουν πλήρως **synced** lyrics — έτοιμα για karaoke/social clips.")
        if unsynced:
            st.warning(f"⚠️ {unsynced} κομμάτια έχουν lyrics αλλά **χωρίς time-sync** — χρειάζονται alignment.")
        if missing:
            st.error(f"❌ {missing} κομμάτια **δεν έχουν καθόλου lyrics** στο Musixmatch — απαιτείται upload.")

        st.divider()
        st.caption("Δείγμα αναλυτικού πίνακα (placeholder):")
        demo_rows = []
        statuses = (["Synced"] * synced) + (["Unsynced"] * unsynced) + (["Missing"] * missing)
        for isrc, status in zip(isrcs, statuses + ["Synced"] * total):
            demo_rows.append({
                "ISRC": isrc,
                "Έγκυρο;": "✅" if validate_isrc(isrc) else "❌",
                "Κατάσταση": status,
                "Γλώσσα": "—",
            })
        st.dataframe(pd.DataFrame(demo_rows), width="stretch", hide_index=True)


# ==========================================================================
# TOOL B: Metadata Health Dashboard
# ==========================================================================
def page_metadata_health():
    st.title("🩺 Metadata Health Dashboard")
    st.caption("Επισκόπηση της ποιότητας του IPI LIST — εντοπίστε writers χωρίς IPI ή PRO.")

    total = None
    problems = None

    # Try to compute REAL metrics from the private IPI LIST; fall back to demo.
    try:
        private_ipi_config = get_private_ipi_config()
        with st.spinner("Ανάλυση IPI LIST..."):
            ipi_file_bytes = fetch_private_ipi_list_bytes(**private_ipi_config)
            ipi_lookup, _ = build_ipi_lookup_from_bytes(ipi_file_bytes)
        total, problems = _scan_ipi_health(ipi_lookup)
        data_is_live = True
    except Exception as e:
        st.info("Χρήση demo δεδομένων (δεν φορτώθηκε το ζωντανό IPI LIST).")
        st.caption(f"Λεπτομέρεια: {e}")
        data_is_live = False
        total = 128
        problems = [
            {"Writer": "Nikos Papadopoulos", "IPI": "—", "PRO": "AEPI", "Πρόβλημα": "Missing IPI"},
            {"Writer": "Maria K.", "IPI": 250123456, "PRO": "—", "Πρόβλημα": "Missing PRO"},
            {"Writer": "John Doe", "IPI": "—", "PRO": "—", "Πρόβλημα": "Missing IPI, Missing PRO"},
        ]

    missing_ipi = sum(1 for p in problems if "Missing IPI" in p["Πρόβλημα"])
    missing_pro = sum(1 for p in problems if "Missing PRO" in p["Πρόβλημα"])

    m1, m2, m3 = st.columns(3)
    m1.metric("Σύνολο Writers στη βάση", total)
    m2.metric("Missing IPIs", missing_ipi, delta=f"-{missing_ipi}" if missing_ipi else "0",
              delta_color="inverse" if missing_ipi else "off")
    m3.metric("Missing PROs", missing_pro, delta=f"-{missing_pro}" if missing_pro else "0",
              delta_color="inverse" if missing_pro else "off")

    st.divider()

    if problems:
        st.warning(f"Βρέθηκαν {len(problems)} writers με ελλιπή metadata — χρειάζονται διόρθωση.")
        st.dataframe(
            pd.DataFrame(problems),
            width="stretch",
            hide_index=True,
            column_config={
                "Writer": st.column_config.TextColumn("Writer"),
                "IPI": st.column_config.TextColumn("IPI"),
                "PRO": st.column_config.TextColumn("PRO"),
                "Πρόβλημα": st.column_config.TextColumn("Πρόβλημα"),
            },
        )
    else:
        st.success("🎉 Όλοι οι writers έχουν πλήρη IPI & PRO metadata!")

    if data_is_live:
        st.caption("Τα παραπάνω προέρχονται από το ζωντανό IPI LIST (private GitHub).")


# ==========================================================================
# TOOL C: Smart Links Generator (real Odesli API)
# ==========================================================================
def page_smart_links():
    st.title("Smart Links Generator")
    st.caption("Δημιουργήστε universal links για όλες τις πλατφόρμες από ένα Spotify URL (μέσω Odesli).")

    url_input = st.text_input(
        "Spotify Track / Album URL",
        placeholder="https://open.spotify.com/track/...",
    )

    if st.button("Δημιουργία Links", type="primary", width="stretch"):
        clean_url = url_input.strip()
        if not clean_url:
            st.warning("Επικολλήστε ένα Spotify URL πρώτα.")
            return

        try:
            with st.spinner("Επικοινωνία με το Odesli API..."):
                data = fetch_odesli_links(clean_url)
        except requests.HTTPError as e:
            st.error(f"Το Odesli API απέρριψε το URL (HTTP {e.response.status_code}). "
                     "Βεβαιωθείτε ότι είναι έγκυρο Spotify link.")
            return
        except Exception as e:
            st.error(f"Αποτυχία επικοινωνίας με το Odesli API: {e}")
            return

        links_by_platform = data.get("linksByPlatform") or {}

        # Optional: nice header with the resolved track/artwork if available.
        page_url = data.get("pageUrl")
        entities = data.get("entitiesByUniqueId") or {}
        unique_id = data.get("entityUniqueId")
        entity = entities.get(unique_id, {}) if unique_id else {}
        title = entity.get("title")
        artist = entity.get("artistName")
        thumb = entity.get("thumbnailUrl")

        st.divider()
        head_col1, head_col2 = st.columns([1, 4], vertical_alignment="center")
        with head_col1:
            if thumb:
                st.image(thumb, width="stretch")
        with head_col2:
            if title:
                st.markdown(f"### {title}")
            if artist:
                st.markdown(f"**{artist}**")
            if page_url:
                st.markdown(f"[Universal song.link σελίδα]({page_url})")

        st.markdown("### Links ανά πλατφόρμα")

        platforms = [
            ("Spotify", "spotify"),
            ("Apple Music", "appleMusic"),
            ("Tidal", "tidal"),
            ("YouTube", "youtube"),
            ("Deezer", "deezer"),
        ]

        found_any = False
        for label, key in platforms:
            platform_data = links_by_platform.get(key)
            if platform_data and platform_data.get("url"):
                found_any = True
                st.markdown(f"**{label}**")
                st.code(platform_data["url"], language=None)  # st.code => built-in copy button
            else:
                st.markdown(f"**{label}** — _δεν βρέθηκε_")

        if not found_any:
            st.info("Δεν επιστράφηκαν links για αυτό το κομμάτι.")


# ==========================================================================
# TOOL D: MusicBrainz Explorer (placeholder / under development)
# ==========================================================================
def page_musicbrainz():
    st.title("MusicBrainz Explorer")

    st.markdown(
        """
        <div class="under-construction">
            <h2>🚧 Under Construction 🚧</h2>
            <p>Deep metadata integration is on the way.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        "Αυτό το εργαλείο θα συνδεθεί απευθείας με το **MusicBrainz** "
        "(JSON/XML API & PostgreSQL δομή) για βαθιά queries metadata:"
    )
    st.markdown(
        "- **Releases & Recordings** — άντληση όλων των εκδόσεων ενός κομματιού\n"
        "- **Legal names & aliases** — επίσημα ονόματα δημιουργών\n"
        "- **ISRCs & barcodes** — cross-reference identifiers\n"
        "- **Label details** — πληροφορίες δισκογραφικής\n"
        "- **Work relationships** — συσχετίσεις έργων, composers, arrangers"
    )
    st.info(
        "Σύντομα θα μπορείτε να κάνετε deep queries κατευθείαν από την "
        "open-source μουσική εγκυκλοπαίδεια, χωρίς χειροκίνητη έρευνα."
    )


# ==========================================================================
# PAGE: Ρυθμίσεις (Settings)
# ==========================================================================
def page_settings(spotify_user):
    st.title("⚙️ Ρυθμίσεις")

    tab_data, tab_system = st.tabs(["Δεδομένα", "Σύστημα"])

    with tab_data:
        st.markdown("### Διαχείριση Cache")
        st.caption(
            "Το IPI LIST και άλλα δεδομένα αποθηκεύονται προσωρινά για ταχύτητα. "
            "Καθαρίστε το cache αν ενημερώσατε το IPI LIST στο GitHub."
        )
        if st.button("🧹 Εκκαθάριση Προσωρινής Μνήμης (Clear Cache)", type="primary"):
            st.cache_data.clear()
            st.toast("Το cache καθαρίστηκε επιτυχώς!", icon="✅")
            st.success("Η προσωρινή μνήμη εκκαθαρίστηκε. Τα δεδομένα θα φορτωθούν ξανά.")

    with tab_system:
        st.markdown("### Πληροφορίες Συστήματος")
        st.text_input("Active Spotify User", value=spotify_user or "—", disabled=True)
        st.text_input("Έκδοση", value=APP_VERSION, disabled=True)
        st.caption("Stay Independent Tool © 2026")


# ==========================================================================
# SIDEBAR NAVIGATION (logged-in) - custom, state-based, highlighted
# ==========================================================================
def _nav_button(label, page_key):
    """Renders a sidebar nav button; highlights the active page via type='primary'."""
    is_active = st.session_state.get("current_page") == page_key
    if st.sidebar.button(
        label,
        width="stretch",
        type="primary" if is_active else "secondary",
        key=f"nav_{page_key}",
    ):
        if not is_active:
            st.session_state.current_page = page_key
            st.rerun()


def render_sidebar(spotify_user):
    with st.sidebar:
        if os.path.exists("StayLogo2.jpg"):
            st.image("StayLogo2.jpg", width="stretch")
        else:
            st.markdown("## 🎵 Stay Independent Tool")

        st.markdown("### Stay Independent Tool\n*Swiss Army Knife*")
        st.success(f"🟢 Συνδεδεμένος: **{spotify_user}**")
        st.divider()

    # --- Top section: Εργαλεία ---
    st.sidebar.markdown("Εργαλεία")
    _nav_button("Γεννήτρια Catalog", "Γεννήτρια Catalog")
    _nav_button("Musixmatch Sync Checker", "Musixmatch Sync Checker")
    _nav_button("Metadata Health", "Metadata Health")
    _nav_button("Smart Links Generator", "Smart Links Generator")
    _nav_button("MusicBrainz Explorer", "MusicBrainz Explorer")

    # --- Spacer to push the System block lower ---
    st.sidebar.markdown("<div style='height: 2.5rem'></div>", unsafe_allow_html=True)
    st.sidebar.divider()

    # --- Bottom section: Σύστημα ---
    st.sidebar.markdown("Σύστημα")
    _nav_button("Ιστορικό & Αρχεία", "Ιστορικό & Αρχεία")
    _nav_button("Ρυθμίσεις", "Ρυθμίσεις")

    if st.sidebar.button("🚪 Αποσύνδεση", width="stretch", key="nav_logout"):
        st.session_state.pop("token_data", None)
        st.session_state.pop("current_page", None)
        st.rerun()

    st.sidebar.divider()
    st.sidebar.caption("Stay Independent Tool © 2026")


# ==========================================================================
# MAIN APPLICATION FLOW
# ==========================================================================
client_id, client_secret, redirect_uri = get_config()

# --- Auth Callback Handling ---
query_params = st.query_params
if "error" in query_params:
    st.error(f"Η σύνδεση με το Spotify απέτυχε: {query_params['error']}")
    st.query_params.clear()
elif "code" in query_params and "token_data" not in st.session_state:
    try:
        token_data = exchange_code_for_token(client_id, client_secret, redirect_uri, query_params["code"])
        st.session_state["token_data"] = token_data
        st.toast("Επιτυχής σύνδεση στο Spotify!", icon="✅")
    except requests.HTTPError as e:
        st.error(f"Σφάλμα κατά την ανταλλαγή του code: {e}")
    st.query_params.clear()
    st.rerun()

token = get_valid_token()

# --- Logged-out: centered landing page, no sidebar ---
if not token:
    render_landing_page(client_id, redirect_uri)
    st.stop()

# --- Logged-in: resolve user, init routing state, render shell ---
try:
    spotify_user = fetch_current_user(token)
except Exception:
    spotify_user = "Άγνωστος Χρήστης"

if "current_page" not in st.session_state:
    st.session_state.current_page = "Γεννήτρια Catalog"

render_sidebar(spotify_user)

# --- State-based router ---
current_page = st.session_state.current_page

if current_page == "Γεννήτρια Catalog":
    page_catalog_generator(token, spotify_user)
elif current_page == "Musixmatch Sync Checker":
    page_musixmatch_checker()
elif current_page == "Metadata Health":
    page_metadata_health()
elif current_page == "Smart Links Generator":
    page_smart_links()
elif current_page == "MusicBrainz Explorer":
    page_musicbrainz()
elif current_page == "Ιστορικό & Αρχεία":
    page_history(token, spotify_user)
elif current_page == "Ρυθμίσεις":
    page_settings(spotify_user)
else:
    # Fallback safety net
    st.session_state.current_page = "Γεννήτρια Catalog"
    st.rerun()
